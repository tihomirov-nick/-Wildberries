import requests
from openpyxl import Workbook, load_workbook


def search(item_id, query):
    max_count = (requests.get(f'https://search.wb.ru/exactmatch/ru/common/v4/search?appType=1&couponsGeo=12,3,18,15,21&curr=rub&dest=-1216601,-337422,-1114902,-1198055&emp=0&lang=ru&locale=ru&pricemarginCoeff=1.0&query={str(query)}&reg=0&regions=68,64,83,4,38,80,33,70,82,86,75,30,69,22,66,31,40,1,48,71&resultset=filters&sort=popular&spp=0')).json()['data']['total']
    counter = 0
    page = 1
    while counter != max_count:
        add_page = f"&page={page}"
        url = f'''https://search.wb.ru/exactmatch/ru/common/v4/search?appType=1&couponsGeo=12,3,18,15,21&curr=rub&dest=-1029256,-102269,-2162196,-1257786&emp=0&lang=ru&locale=ru{add_page}&pricemarginCoeff=1.0&query={str(query)}&reg=0&regions=68,64,83,4,38,80,33,70,82,86,75,30,69,22,66,31,40,1,48,71&resultset=catalog&sort=popular&spp=0&suppressSpellcheck=false'''
        request = requests.get(url).json()
        for i in range(100):
            if str(item_id) == str(request['data']['products'][i]['id']):
                return counter + 1
            counter += 1
        page += 1


def find_info(item_id):

    info_url = f'''https://card.wb.ru/cards/detail?spp=0&regions=68,64,83,4,38,80,33,70,82,86,75,30,69,22,66,31,40,1,48,71&pricemarginCoeff=1.0&reg=0&appType=1&emp=0&locale=ru&lang=ru&curr=rub&couponsGeo=12,7,3,6,5,18,21&dest=-1216601,-337422,-1114902,-1198055&nm={item_id}'''
    info = requests.get(url=info_url)
    rating = info.json()['data']['products'][0]['rating']
    feedbacks = info.json()['data']['products'][0]['feedbacks']
    brand = info.json()['data']['products'][0]['brand']
    try:
        price = int(info.json()['data']['products'][0]['extended']['basicPriceU']) / 100
    except:
        price = int(info.json()['data']['products'][0]['priceU']) / 100

    card_url = f'''https://wbx-content-v2.wbstatic.net/ru/{item_id}.json'''
    add_info = requests.get(url=card_url)
    options = add_info.json()['options']

    weight = "Не указан"

    for i in range(len(options) - 1):
        if str(options[i]['ao_id']) == "89008":
            weight = str(options[i]['value'])
            break

    info_url = f"https://www.wildberries.ru/catalog/{item_id}/detail.aspx?targetUrl=XS"

    return [f"{info_url}", f"{rating}", f"{feedbacks}", f"{brand}", f"{int(price)}р.", f"{weight}"]


def create_exel():
    wb = load_workbook(filename="data.xlsx")
    sheet = wb["Sheet"]
    sheet.append(["", "Рейтинг", "Отзывы", "Бренд", "Цена", "Вес"])
    wb.save("data.xlsx")


def add_to_exel(line):
    wb = load_workbook(filename="data.xlsx")
    sheet = wb["Sheet"]
    ret = line
    sheet.append(ret)
    wb.save("data.xlsx")


def top(query):
    url = f'''https://search.wb.ru/exactmatch/ru/common/v4/search?appType=1&couponsGeo=12,3,18,15,21&curr=rub&dest=-1216601,-337422,-1114902,-1198055&emp=0&lang=ru&locale=ru&pricemarginCoeff=1.0&query={str(query)}&reg=0&regions=68,64,83,4,38,80,33,70,82,86,75,30,69,22,66,31,40,1,48,71&resultset=catalog&sort=popular&spp=0&suppressSpellcheck=false'''
    response = requests.get(url=url)
    top_10_items = []
    for i in range(10):
        top_10_items.append(response.json()['data']['products'][i]['id'])
    return top_10_items


def main_cycle(query, item_id):
    top_10_items = top(query)
    for i in range(len(top_10_items)):
        add_to_exel(find_info(top_10_items[i]))
    add_to_exel(find_info(item_id) + [str(search(item_id, query))] + [f'''"{query}"'''])


if __name__ == '__main__':
    Workbook().save("data.xlsx")
    create_exel()

    f = open('input.txt', encoding='utf-8', mode='r')
    arr = (str(f.read())).split("\n")
    for i in range(len(arr)):
        main_cycle(str(arr[i]).split("-")[0], str(arr[i]).split("-")[1])
